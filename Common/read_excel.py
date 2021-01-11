# -*- coding: utf-8 -*-
import re
import pandas as pd
import logging, logging.config
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
from TestCase.TestErp.set_var_erp import GetData
# from TestCase.TestJht.set_var_jht import GetData1
from Common.ProjectPath.file_path import log_config_path

logging.config.fileConfig(log_config_path)


class DoExcel(object):

    @staticmethod
    def pandas_data(file_name, sheet_name):
        df = pd.read_excel(file_name, sheet_name=sheet_name)
        test_data = []
        for i in df.index.values:  # 获取行号的索引，并对其进行遍历：
            row_data = df.loc[i, list(df.head(0))].to_dict()
            test_data.append(row_data)
        logging.info("最终获取到的数据是：{0}".format(test_data))
        for x in test_data:
            pass
            # print(x)
        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, res, result, pass_time):
        # 测试响应写回excel文件
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]

        font_pass = Font(size=11, bold=True, color="00FF00")
        font_false = Font(size=11, bold=True, color="FF0000")

        sheet.cell(i, 8).value = res  # response实际响应结果
        if result == '通过':
            sheet.cell(i, 9).font = font_pass
        else:
            sheet.cell(i, 9).font = font_false
        sheet.cell(i, 9).value = result  # 记录Pass/Fail
        sheet.cell(i, 7).value = pass_time  # 记录Pass/Fail
        sheet.row_dimensions[i].height = 30  # 设置sheet表行高20
        wb.save(file_name)

    @staticmethod
    def re_find(res_text='', find_col=''):
        try:
            if find_col == 'VIEWSTATE':
                VIEWSTATE = \
                    re.findall(r'<input type="hidden" name="__VIEWSTATE" id="__VIEWSTATE" value="(.*?)" />', res_text,
                               re.I)[0]
                return VIEWSTATE
            elif find_col == 'VIEWSTATEGENERATOR':
                VIEWSTATEGENERATOR = \
                    re.findall(r'name="__VIEWSTATEGENERATOR" id="__VIEWSTATEGENERATOR" value="(.*?)" />', res_text,
                               re.I)[0]
                return VIEWSTATEGENERATOR
            elif find_col == 'EVENTVALIDATION':
                EVENTVALIDATION = \
                    re.findall(r'name="__EVENTVALIDATION" id="__EVENTVALIDATION" value="(.*?)" />', res_text, re.I)[0]
                return EVENTVALIDATION
            else:
                re_data = re.findall(find_col, str(res_text), re.I)[0]
                if re_data:
                    logging.info(f'正则匹配：{re_data}')
                return re_data
        except:
            return None

    @staticmethod
    def re_set_data(res_url='', res_text='', re_pat='', method='post'):
        pass

    def re_replace(self, data):
        # 正则数据替换
        pat = r'\${(.*?)}'
        while re.search(pat, data):
            key = re.search(pat, data).group(0)
            value = re.search(pat, data).group(1)
            data = data.replace(key, getattr(GetData, value))
        logging.info('请求数据：' + data)
        return data

    def re_replaces(self, class_data, data):
        # 正则数据替换
        pat = r'\${(.*?)}'
        while re.search(pat, data):
            key = re.search(pat, data).group(0)
            value = re.search(pat, data).group(1)
            data = data.replace(key, getattr(class_data, value))
        logging.info('请求数据：' + data)
        return data

    def pd_read_case(self, file_name, mode):
        logging.info('测试的模块:' + str(mode))
        test_data = []
        for key in mode:
            sheet = pd.read_excel(file_name, sheet_name=key)
            if mode[key] == 'all':
                # print('指定运行的模块运行所有:{0}'.format(mode.keys()))
                for i in sheet.index.values:
                    row_data = sheet.loc[i, list(sheet.head(0))].to_dict()
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
            elif type(mode[key]) is list:
                for case_id in mode[key]:
                    # print('指定运行的case_id:', case_id)
                    row_data = sheet.loc[case_id - 1, list(sheet.head(0))].to_dict()
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
        for i in test_data:
            # logging.info('测试数据:' + str(i))
            pass
        logging.info('执行case个数: ' + str((len(test_data))))
        return test_data

    @staticmethod  # 正则二次开发及替换
    def common_replace(request_data, return_data, url, pat=r'io_id\\":(.*?),', dict_key='io_id'):
        try:
            if request_data['请求方法'] == 'post' and request_data['路径'] == url:
                try:
                    io_id = DoExcel.re_find(return_data, pat)
                    if io_id:
                        x = getattr(GetData, 'dict_var')
                        x[dict_key] = io_id
                        logging.info(x)
                except:
                    pass
        except:
            logging.info('.......数据有问题......')


if __name__ == '__main__':
    DoExcel.pandas_data(file_name=r'/TestCase/TestErp/test_erp.xlsx',
                        sheet_name='直接发货')
    pass
